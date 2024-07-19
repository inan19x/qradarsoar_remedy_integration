import resilient
import re
import json 
import logging 
import smtplib
import ssl
import openpyxl
import os
import resilient_lib
import time
import shutil
import requests

from datetime import datetime
from xml.dom import minidom

from circuits.core.handlers import handler
from resilient_circuits.actions_component import ResilientComponent, ActionMessage

logger = logging.getLogger(__name__)

class RemedyTicketProcessor(ResilientComponent):
	# Subscribe to the Action Module message destination named 'remedy'
	channel = 'actions.remedy'


	def create_dict_from_2_columns(self, workbook, worksheet, key_col, value_col):
		mapping = {}

		sheet_obj = workbook[worksheet]
		m_row = sheet_obj.max_row 
		
		keys = []
		values = []

		for i in range(2, m_row + 1): 
			cell_obj = sheet_obj.cell(row = i, column = key_col) 
			keys.append(cell_obj.value) 

		for i in range(2, m_row + 1): 
			cell_obj = sheet_obj.cell(row = i, column = value_col) 
			values.append(cell_obj.value) 

		counter = len(keys)
		count = 0

		while count < counter:
			mapping[keys[count]] = values[count]
			count += 1

		return mapping


	def get_artifacts(self, client, incident_id):
		uri = "/incidents/{}/artifacts".format(incident_id)
		artifacts = client.get(uri)

		artifacts_list = []

		for artifact in artifacts:
			if artifact['relating']:
				uri = '/artifact_types/{}'.format(artifact['type'])
				type_info = client.get(uri)
				type_artifact = type_info['name']

				if artifact['type'] == 1 and artifact['properties'] is not None:
					try:
						type_artifact = type_info['name'] + ':' + artifact['properties'][0]['name']
					finally:
						type_artifact = type_info['name']

				value 			= artifact['value'].encode(encoding="utf-8",errors="strict").strip()
				description_ 	= artifact['description'].encode(encoding="utf-8",errors="strict").strip() if artifact['description'] is not None else ''
				artifacts_list.append("{0}		: {1} ({2})".format(type_artifact, value, description_))

		return artifacts_list


	def login(self):
		url = "https://<remedy_host>/api/jwt/login"
		headers = {
			"Content-Type": "application/x-www-form-urlencoded"
		}
		data = {
			"username": "<remedy_user>",
			"password": "<remedy_pass>",
			"authString": "authentication"
		}

		response = requests.post(url, headers=headers, data=data, verify=False)
		if response.status_code == 200:
			return response.text 
		else:
			return None


	def submit_ticket(self, jwt_token, json_data):
		url = "https://<remedy_host>/api/arsys/v1/entry/HPD:IncidentInterface_Create?fields=values(Incident Number)"
		content_length = len(json.dumps(json_data))
		headers = {
			"Host": "<remedy_host>",
			"Content-Type": "application/json",
			"Authorization": "AR-JWT {0}".format(jwt_token)	
		}
		headers["Content-Length"] = str(content_length)
		payload = json.dumps(json_data).strip('"')
		payload = payload.replace('\\"','"').replace('\\n', '\n').replace('\\t', '\t')
	
		response = requests.post(url, headers=headers, data=payload, verify=False)
		return response.status_code, response.text


	def get_entry_id(self, jwt_token, incident_number):
		url = "https://<remedy_host>/api/arsys/v1/entry/HPD:IncidentInterface?q='Incident Number'=\"{}\"".format(incident_number)
		headers = {
			"Authorization": "AR-JWT {0}".format(jwt_token),
			"Content-Type": "application/json"
		}

		response = requests.get(url, headers=headers, verify=False)
		if response.status_code == 200:
			entries = response.json().get('entries', [{}])[0].get('values' , {}).get('Entry ID')
			return entries 
		else:
			return None


	def update_ticket(self, jwt_token, entry_id):
		url = "https://<remedy_host>/api/arsys/v1/entry/HPD:IncidentInterface/{0}|{0}".format(entry_id)
		headers = {
			"Authorization": "AR-JWT {0}".format(jwt_token),
			"Content-Type": "application/json"
		}
		data = {
			"values": {
				"Status": "Resolved",
				"Status_Reason": "No Further Action Required",
				"Resolution": "Automatically updated to resolved by QRadar SOAR"
			}
		}
		
		response = requests.put(url, headers=headers, json=data, verify=False)
		return response.text 


	# first escalation
	@handler('create_ticket_in_remedy')
	def _create_ticket_in_remedy_handler_function(self, event, headers, *args, **kwargs):
		incident = event.message['incident']

		#Init variables
		incident_id = incident['id']
		qradar_id = incident['properties']['qradar_id']
		incident_name = incident['name'].encode(encoding="utf-8",errors="strict").strip()
		description = incident['description'].encode(encoding="utf-8",errors="strict").strip()
		description_esc = description.replace('"','\\"')
		offense_source = incident['properties']['offense_source'].encode(encoding="utf-8",errors="strict").strip()
		incident_types = incident['incident_type_ids']
		date_occured = datetime.fromtimestamp(float(incident['discovered_date'])/1000).strftime('%e %b %Y, %H:%M')
		remedy_group_id = event.message['properties']['remedy_group']
		remedy_group_name = event.message['type_info']['actioninvocation']['fields']['remedy_group']['values'][str(remedy_group_id)]['label']

		#Configure SOAR connection
		parser = resilient.ArgumentParser(config_file=resilient.get_config_file())
		opts = parser.parse_args()
		client = resilient.get_client(opts)
		client.context_header = headers['Co3ContextToken']

		#Get SOAR /types properties
		uri = "/types"
		types = client.get(uri)

		severity_code = {}
		impact_list = {}
		incident_type_ids = {}

		inc_type_ids = types['incident']['fields']['incident_type_ids']['values']
		for inc_t_id in inc_type_ids:
			incident_type_ids[inc_t_id['value']] = inc_t_id['label']

		severity_code[None] = 'None'
		impact_list[None] = 'None'
		
		the = types['incident']['fields']['severity_code']['values']

		for t in the:
			severity_code[t['value']] = t['label']

		the = types['incident']['fields']['impact']['values']

		for t in the:
			impact_list[t['value']] = t['label']

		incident_severity = incident['severity_code']
		incident_impact = incident['properties']['impact']

		#Mapping urgency in Remedy
		remedy_urgency = 'N/A'
		if severity_code[incident_severity] == 'Critical':
			remedy_urgency = '1-Urgent'
		elif severity_code[incident_severity] == 'High':
			remedy_urgency = '2-High'
		elif severity_code[incident_severity] == 'Medium':
			remedy_urgency = '3-Medium'
		else:
			remedy_urgency = '4-Low'

		# Populate assigned group as per defined guidelines in Excel file
		wb_obj = openpyxl.load_workbook("SOC_Playbook_Escalation.xlsx") 
		
        # Worksheet in Excel file to be used
		worksheet = 'Remedy Resolver Group'

		# Populate the ticket information
		mapping_team_vs_group = self.create_dict_from_2_columns(wb_obj, worksheet, 1, 3)
		mapping_team_vs_group_id = self.create_dict_from_2_columns(wb_obj, worksheet, 1, 4)
		mapping_team_vs_group_org = self.create_dict_from_2_columns(wb_obj, worksheet, 1, 2)
		mapping_team_vs_assignee = self.create_dict_from_2_columns(wb_obj, worksheet, 1, 5)
		artifacts_list = self.get_artifacts(client, incident_id)
		Detailed_Description = "Kindly follow the recommendation in email sent at {0} with subject: Incident {1}".format(date_occured, incident_id)

		#Convert to regular string
		impact_list[incident_impact] = str(impact_list[incident_impact])
		mapping_team_vs_group_org[remedy_group_name] = str(mapping_team_vs_group_org[remedy_group_name])
		mapping_team_vs_group_id[remedy_group_name] = str(mapping_team_vs_group_id[remedy_group_name])
		mapping_team_vs_group[remedy_group_name] = str(mapping_team_vs_group[remedy_group_name])
		mapping_team_vs_assignee[remedy_group_name] = str(mapping_team_vs_assignee[remedy_group_name])
		incident_tier_3 = str(incident_type_ids[incident_types[0]].upper())

		json_data = r'''
{
    "values": {
        "Person ID": "<remedy_PersonId>",
        "First_Name": "<remedy_First_Name>",
        "Last_Name": "<remedy_Last_Name>",
        "Service_Type": "User Service Request",
        "Status": "Assigned",
        "Impact": "%s",
        "Urgency": "%s",
        "Description": "Security incident escalated to resolver group",
        "Reported Source": "Direct Input",
        "Product Categorization Tier 1": "SECURITY SYSTEM",
        "Product Categorization Tier 2": "SOC",
        "Product Categorization Tier 3": "QRADAR SOAR",
        "Categorization Tier 1": "REQUEST",
        "Categorization Tier 2": "SECURITY INCIDENT (SI)",
        "Categorization Tier 3": "%s",
        "z1D_Action": "CREATE",
        "Assigned Support Company": "ACME Ltd.",
        "Assigned Support Organization": "%s",
        "Assigned Group ID": "%s",
        "Assigned Group": "%s",
        "Assignee": "%s",
        "Resolution": "-",
        "Detailed_Decription": "%s",
        "ServiceCI_ReconID" : "<remedy_ServiceCI>",
        "Status_Reason": "-"
    }
}
		''' % (impact_list[incident_impact], remedy_urgency, incident_tier_3, mapping_team_vs_group_org[remedy_group_name],
       		mapping_team_vs_group_id[remedy_group_name], mapping_team_vs_group[remedy_group_name],
      	 	mapping_team_vs_assignee[remedy_group_name], Detailed_Description)

		#Get Remedy login jwt token
		jwt_token = self.login()

		#Create Remedy ticket
		status_resp = 0
		status_resp, new_incident = self.submit_ticket(jwt_token,json_data)

		if status_resp == 201:
			remedy_data = json.loads(new_incident)
			new_ticket_id = remedy_data['values']['Incident Number']
			logger.info('Creating ticket success! Ticket ID: {0}'.format(new_ticket_id))
			
			#Get SOAR incident properties
			uri = '/incidents/{0}'.format(incident_id)
			incident = client.get(uri)

			#Update remedy ticket informations in SOAR
			incident['properties']['remedy_ticket_id'] = new_ticket_id
			incident['properties']['remedy_ticket_status'] = "Assigned"
			incident['properties']['assigned_group_remedy'] = mapping_team_vs_group[remedy_group_name]
			client.put(uri, incident)

		else: 
			logger.info('Creating ticket failed! Message: {0} - {1}'.format(status_resp, new_incident))


	#Close remedy ticket
	@handler('close_ticket_in_remedy')
	def _close_ticket_in_remedy_handler_function(self, event, headers, *args, **kwargs):
		incident = event.message['incident']

		#Init variables
		incident_id = incident['id']
		
		#Configure SOAR Connection
		parser = resilient.ArgumentParser(config_file=resilient.get_config_file())
		opts = parser.parse_args()
		client = resilient.get_client(opts)
		client.context_header = headers['Co3ContextToken']

		#Get SOAR incident properties
		uri = '/incidents/{0}'.format(incident_id)
		incident = client.get(uri)

		remedy_ticket_id = incident['properties']['remedy_ticket_id']

		#Get Remedy login jwt token
		jwt_token = self.login()
	
		#If logged in, update Remedy ticket properties
		if jwt_token:
			entry_id = self.get_entry_id(jwt_token, remedy_ticket_id)
			if entry_id:
				remedy_ticket_update = self.update_ticket(jwt_token, entry_id)
				logger.info('Remedy ticket ID {0} closed!'.format(remedy_ticket_id))
				
				#Get SOAR incident properties
				uri = '/incidents/{0}'.format(incident_id)
				incident = client.get(uri)

				#Update Remedy ticket informations in SOAR
				incident['properties']['remedy_ticket_status'] = "Resolved"
				client.put(uri, incident)
